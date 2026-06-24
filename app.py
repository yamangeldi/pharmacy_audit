import os
import io
import pandas as pd
import requests
import openpyxl
import streamlit as st
from roboflow import Roboflow
from openpyxl.styles import PatternFill, Font

# ==========================================
# 1. СЛОВАРЬ КРАСИВЫХ НАЗВАНИЙ ПРЕПАРАТОВ
# ==========================================
PRODUCT_NAMES = {
    "cardiom_150_100": "Кардиомагнил 150мг",
    "ent_2_10": "Энтерожермина 2мл №10",
    "ent_2_12": "Энтерожермина 2мл №12",
    "ent_4_10": "Энтерожермина 4мл №10",
    "magneb6_60": "Магне В6 №60",
    "nospa_40_24": "Но-шпа 40мг №24",
    "snup_01_15": "Снуп 0.1% 15мл",
    "zodac_10_30": "Зодак 10мг №30",
    "las_15_5_100": "Лазолван 15мг"
}

st.set_page_config(page_title="Аудит Выкладки", layout="centered")

st.image("logo.png", width=250)
st.title("💊 Автоматический аудит выкладки")
st.write("Загрузите файл матрицы (план) и файл с фотоотчетом (ссылками), чтобы нейросеть проверила наличие препаратов.")

st.subheader("1. Загрузка данных")
matrix_file = st.file_uploader("Загрузите матрицу (matrix.xlsx)", type=["xlsx"])
report_file = st.file_uploader("Загрузите фотоотчет (Oson...xlsm или xlsx)", type=["xlsx", "xlsm"])

api_key_input = st.text_input("Введите ваш API-ключ Roboflow", type="password")

if st.button("🚀 Запустить проверку", type="primary"):
    if not matrix_file or not report_file or not api_key_input:
        st.error("Пожалуйста, загрузите оба файла и введите API-ключ!")
    else:
        try:
            st.info("Подключение к нейросети...")
            rf = Roboflow(api_key=api_key_input)
            project = rf.workspace().project("uz_ir_pharmacy")
            model = project.version(1).model # Поменяй на нужную версию, если обновил модель
            
            st.info("Чтение планов матрицы...")
            df_matrix = pd.read_excel(matrix_file)
            df_matrix = df_matrix.dropna(subset=['roboflow_name'])

            matrix_plans = {}
            all_unique_products = set()
            for index, row in df_matrix.iterrows():
                shelf_name = str(row['Column Name']).strip()
                product = str(row['roboflow_name']).strip()
                target = int(row['Target packs'])
                
                if shelf_name not in matrix_plans:
                    matrix_plans[shelf_name] = {}
                matrix_plans[shelf_name][product] = target
                all_unique_products.add(product)

            all_unique_products = sorted(list(all_unique_products))

            st.info("Открытие фотоотчета...")
            wb = openpyxl.load_workbook(report_file, data_only=True)
            ws = wb.active
            
            max_rows = ws.max_row + 1 
            master_results = []
            
            progress_bar = st.progress(0)
            status_text = st.empty()

            total_items = max_rows - 2
            if total_items <= 0:
                st.warning("В загруженном отчете нет данных для проверки.")
                st.stop()

            # ==========================================
            # 2. ОСНОВНОЙ РАСЧЕТ И ЛОГИКА КОЛОНОК
            # ==========================================
            for i, row_idx in enumerate(range(2, max_rows)):
                status_text.text(f"Анализ строки {row_idx} (аптека {ws.cell(row=row_idx, column=3).value})...")
                progress_bar.progress((i + 1) / total_items)
                
                visit_id = ws.cell(row=row_idx, column=1).value
                if visit_id is None:
                    continue

                visit_date = ws.cell(row=row_idx, column=2).value
                pharmacy_id = ws.cell(row=row_idx, column=3).value
                shelf_name = str(ws.cell(row=row_idx, column=7).value).strip()
                
                cell_url = ws.cell(row=row_idx, column=9)
                image_url = cell_url.hyperlink.target if hasattr(cell_url, 'hyperlink') and cell_url.hyperlink else str(cell_url.value)

                if not image_url or not str(image_url).startswith("http"):
                    continue

                row_data = {
                    "ID Визита": visit_id,
                    "Дата": visit_date,
                    "ID Аптеки": pharmacy_id,
                    "Название полки": shelf_name,
                    "Ссылка на фото": image_url
                }

                temp_filename = f"temp_visit_{row_idx}.jpg"
                try:
                    response = requests.get(image_url, timeout=15)
                    if response.status_code == 200:
                        with open(temp_filename, "wb") as f:
                            f.write(response.content)
                            
                    prediction = model.predict(temp_filename, confidence=40, overlap=30).json()
                    found_items = prediction.get("predictions", [])

                    shelf_fact = {}
                    for item in found_items:
                        cls = item["class"]
                        shelf_fact[cls] = shelf_fact.get(cls, 0) + 1

                    current_plan = matrix_plans.get(shelf_name, {})
                    
                    sum_of_percentages = 0
                    planned_items_count = 0
                    total_plan_packs = 0
                    total_fact_packs = 0

                    for product in all_unique_products:
                        readable_name = PRODUCT_NAMES.get(product, product)
                        
                        col_pct = f"% {readable_name}"
                        col_fact = f"Шт. {readable_name}"

                        target = current_plan.get(product, 0)

                        if target > 0:
                            fact = shelf_fact.get(product, 0)
                            pct = min((fact / target) * 100, 100) 
                            
                            row_data[col_pct] = round(pct, 1)
                            row_data[col_fact] = fact 
                            
                            sum_of_percentages += pct
                            planned_items_count += 1
                            total_plan_packs += target
                            total_fact_packs += fact
                        else:
                            row_data[col_pct] = "Нет плана"
                            row_data[col_fact] = "Нет плана"

                    if planned_items_count > 0:
                        row_data["ИТОГО ПО ПОЛКЕ (%)"] = round(sum_of_percentages / planned_items_count, 1)
                        row_data["Факт полки (шт)"] = total_fact_packs
                        row_data["План полки (шт)"] = total_plan_packs
                    else:
                        row_data["ИТОГО ПО ПОЛКЕ (%)"] = "Нет плана"
                        row_data["Факт полки (шт)"] = "-"
                        row_data["План полки (шт)"] = "-"

                    master_results.append(row_data)

                except Exception as e:
                    print(f"Ошибка в строке {row_idx}: {e}")
                finally:
                    if os.path.exists(temp_filename):
                        os.remove(temp_filename)

            # ==========================================
            # 3. СБОРКА И КОМБИНИРОВАННАЯ РАСКРАСКА
            # ==========================================
            if master_results:
                base_cols = ["ID Визита", "Дата", "ID Аптеки", "Название полки", "Ссылка на фото", 
                             "ИТОГО ПО ПОЛКЕ (%)", "Факт полки (шт)", "План полки (шт)"]
                
                product_cols = []
                for p in all_unique_products:
                    readable_name = PRODUCT_NAMES.get(p, p)
                    product_cols.extend([f"% {readable_name}", f"Шт. {readable_name}"])
                
                df_final = pd.DataFrame(master_results)[base_cols + product_cols]
                
                st.success("✅ Анализ завершен!")
                st.dataframe(df_final)
                
                buffer = io.BytesIO()
                
                # Цвета для ЗАГОЛОВКОВ (Строка 1)
                color_totals_header = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Светло-желтый для ИТОГО
                product_colors_list = [
                    PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"), # Светло-зеленый
                    PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"), # Светло-синий
                    PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"), # Персиковый
                    PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid"), # Светло-фиолетовый
                    PatternFill(start_color="FDCEE8", end_color="FDCEE8", fill_type="solid"), # Светло-розовый
                    PatternFill(start_color="D0F0C0", end_color="D0F0C0", fill_type="solid")  # Чайный зеленый
                ]

                # Цвета для ДАННЫХ (Светофор ниже 1 строки)
                fill_green = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid") # Нежно-зеленый (100%)
                fill_red = PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid")   # Нежно-розовый (<100%)
                fill_grey = PatternFill(start_color="F1F3F4", end_color="F1F3F4", fill_type="solid")  # Светло-серый (Not in Plan)

                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    df_final.to_excel(writer, index=False, sheet_name='Аудит')
                    ws_out = writer.sheets['Аудит']
                    
                    headers = [cell.value for cell in ws_out[1]]
                    
                    # КРАСИМ ЗАГОЛОВКИ (Только строка 1)
                    for col_idx, header in enumerate(headers, 1):
                        # Сделаем текст заголовков жирным для красоты
                        ws_out.cell(row=1, column=col_idx).font = Font(bold=True)
                        
                        if header in ["ИТОГО ПО ПОЛКЕ (%)", "Факт полки (шт)", "План полки (шт)"]:
                            ws_out.cell(row=1, column=col_idx).fill = color_totals_header
                        else:
                            for p_idx, p in enumerate(all_unique_products):
                                readable_name = PRODUCT_NAMES.get(p, p)
                                if header == f"% {readable_name}" or header == f"Шт. {readable_name}":
                                    color_to_use = product_colors_list[p_idx % len(product_colors_list)]
                                    ws_out.cell(row=1, column=col_idx).fill = color_to_use
                                    break
                    
                    # КРАСИМ ДАННЫХ (Строки со 2-й и ниже) — Только проблемные зоны (<100%)
                    for row_idx in range(2, len(df_final) + 2):
                        for col_idx, header in enumerate(headers, 1):
                            cell_value = ws_out.cell(row=row_idx, column=col_idx).value
                            
                            if header == "ИТОГО ПО ПОЛКЕ (%)" or str(header).startswith("%"):
                                # Красим только если это число и оно меньше 100
                                if isinstance(cell_value, (int, float)) and cell_value < 100:
                                    ws_out.cell(row=row_idx, column=col_idx).fill = fill_red

                st.download_button(
                    label="📥 Скачать итоговый отчет",
                    data=buffer.getvalue(),
                    file_name="Итоговый_Отчет.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.warning("Данные не найдены.")
                
        except Exception as e:
            st.error(f"Произошла ошибка: {e}")